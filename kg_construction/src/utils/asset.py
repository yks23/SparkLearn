import pandas as pd
import networkx as nx
from openpyxl import load_workbook
from src.utils.id_operation import graph_structure, GraphStructureType
import re


def section_name_transform(name: str):
    re_expression =[r"(# 第\d+部分)", r"(## 第\d+章)", r"(## \d+\.\d+)"]
    for re_exp in re_expression:
        name = re.sub(re_exp, "", name)
    name=name.replace(" ", "")
    name=name.replace("#", "")
    name=name.replace("\n", "")
    return name

def export_knowledge_graph_to_excel(nodes, edges, start_id: int = None, output_path="knowledge_graph.xlsx"):
    """
    将知识图谱导出为 Excel 文件，符合特定的层级结构格式，并提取非树边写入“节点关系”列。

    参数:
    - nodes: 列表，包含节点信息，每个元素为 {"id": 节点ID, "name": 节点名称}
    - edges: 列表，包含关系信息，每个元素为 {"source": 起始节点ID, "target": 目标节点ID, "relation": 关系描述}
    - start_id: 可选，指定 BFS 的起始节点 ID，如果为空，则自动选择根节点
    - output_path: 导出 Excel 文件的路径
    """

    # 1️⃣ 构建有向图
    G = nx.DiGraph()

    # 添加节点
    node_dict = {node["id"]: node["name"] for node in nodes}
    for node in nodes:
        G.add_node(node["id"], name=node["name"])

    # 添加关系
    for edge in edges:
        G.add_edge(edge["source"], edge["target"], relation=edge["relation"])

    # 2️⃣ 找到根节点（即没有入边的节点）
    if start_id:
        root_nodes = [start_id]
    else:
        root_nodes = [node for node in G.nodes if G.in_degree(node) == 0]

    if not root_nodes:
        raise ValueError("没有找到根节点，可能是图结构错误（存在循环依赖）。")

    # 3️⃣ 进行 BFS 并构建树结构，同时记录非树边
    all_paths = []
    tree_edges = set()
    non_tree_edges = {}
    nodes=[]
    def bfs_tree(start_node):
        """广度优先搜索，构建 BFS 树，并存储层级路径，同时记录非树边"""
        queue = [(start_node, [start_node])]  # (当前节点, BFS路径)
        visited_nodes = set()  # 记录已访问的节点
        visited_nodes.add(start_node)

        while queue:
            node, path = queue.pop(0)
            children = list(G.successors(node))
            all_paths.append([node_dict[n] for n in path])  # 转换 ID 为 名称
            print("explore node",len(all_paths))
            nodes.append(path[-1])  # 记录节点名称
            for child in children:
                if child in visited_nodes:  # 避免循环依赖
                    # 记录非树边
                    if node in non_tree_edges:
                        non_tree_edges[node].append(child)
                    else:
                        non_tree_edges[node] = [child]
                    continue

                queue.append((child, path + [child]))  # 添加子节点路径
                visited_nodes.add(child)
                tree_edges.add((node, child))  # 标记树边

    # 遍历所有可能的根节点，执行 BFS
    for root in root_nodes:
        bfs_tree(root)

    # 4️⃣ 填充层级，确保所有路径达到 10 级
    max_depth = 10
    
    for path in all_paths:
        path.extend([""] * (max_depth - len(path)))  # 用空字符串填充至 7 级

    # 5️⃣ 转换为 DataFrame
    column_names = [
        "根节点",
        "二级知识点",
        "三级知识点",
        "四级知识点",
        "五级知识点",
        "六级知识点",
        "七级知识点",
        "八级知识点",
        "九级知识点",
        "十级知识点",
    ]
    df = pd.DataFrame(all_paths, columns=column_names)
    # 6️⃣ 生成“节点关系”列
    node_relationships = []
    for node_id in nodes:
        if node_id and node_id in non_tree_edges:
            related_nodes = [f"相关:{node_dict[n]}" for n in non_tree_edges[node_id]]
            node_relationships.append(",".join(related_nodes))
        else:
            node_relationships.append("")

    df["知识说明"] = ""
    df["知识点标签"] = ""
    df["节点关系"] = node_relationships  # 插入提取的非树边关系
    df["思政元素"] = ""
    df["设计方式"] = ""

    # 7️⃣ 先保存 DataFrame，保证列名在第二行
    df.to_excel(output_path, index=False)

    # 8️⃣ 读取 Excel 并插入模板说明
    wb = load_workbook(output_path)
    ws = wb.active

    # 定义模板说明文本
    template_text = [
        "模板使用说明：",
        "1. A列至F列对应根节点到七级知识点",
        "   填写示例如下，中间知识点不可为空",
        "   （1）需要新增一个【四级知识点】，需填写【根节点】-【二级知识点】-【三级知识点】-【四级知识点】",
        "   （2）需要新增一个【六级知识点】，需填写【根节点】-【二级知识点】-【三级知识点】-【四级知识点】-【五级知识点】-【六级知识点】",
        "2. 知识点标签：标签支持自定义，填写结构为标签所属分类:标签名称，支持填入多个，以,(英)隔开即可。例：重难点:难点,认知维度:记忆",
        "3. 节点关系：节点关系支持自定义，填写结构为关系名称:关联知识点名称，支持填入多个，以,(英)隔开即可。其中仅“先修”，“后修”及“相关”关系会在图谱内展示对应关系。",
        "4. 思政元素：若知识点为思政知识点，则可填写对应的思政元素，例：道德伦理教育、爱国主义教育、法治教育、社会责任教育、科学文化教育等。",
        "5. 设计方式：如标记思政元素后，可填写思政元素设计方式。如果无思政元素则无需填写。每种思政元素只能对应一种设计方式，如同一思政元素存在多种设计方式，系统会保留最后一种。",
        "6. 请不要删除此行，也不要删除模板中的任何列",
    ]

    # 9️⃣ 插入模板说明在第一行（不覆盖列名）
    ws.insert_rows(1)  # 插入一行，把列名下移
    for i, text in enumerate(template_text, start=1):
        ws.cell(row=1, column=i, value=text)

    # 🔟 保存最终 Excel
    wb.save(output_path)

    print(f"知识图谱数据已成功导出至 {output_path}")


def save_to_excel(data_path: str, output_path: str,start_id:int=None):
    [nodes, edges] = graph_structure(
        [GraphStructureType.all_node, GraphStructureType.all_relation],
        return_type="dict",
        cache_path=data_path,
    )
    nodes = [
        {"id": node["id"], "name": section_name_transform(node["title"])}
        for node in nodes
    ]
    edges = [
        {
            "source": edge["source_id"],
            "target": edge["target_id"],
            "relation": edge["type"],
        }
        for edge in edges
    ]
    export_knowledge_graph_to_excel(nodes, edges, output_path=output_path,start_id=start_id)
    print("导出成功")


def main():
    nodes = [
    {"id": 1, "name": "电磁学"},
    {"id": 2, "name": "第1章 静电场"},
    {"id": 3, "name": "第2章 电介质"},
    {"id": 4, "name": "第3章 电磁场"},
    {"id": 5, "name": "电荷"},
    {"id": 6, "name": "库仑定律"},
    {"id": 7, "name": "高斯定理"},
    {"id": 8, "name": "电容器"},
    {"id": 9, "name": "磁质分类"},
    {"id": 10, "name": "法拉第效应"},
    {"id": 11, "name": "安培定理"},
    {"id": 12, "name": "麦克斯韦方程组"},
]
    edges = [
    {"source": 1, "target": 2, "relation": "has_chapter"},
    {"source": 1, "target": 3, "relation": "has_chapter"},
    {"source": 1, "target": 4, "relation": "has_chapter"},
    
    {"source": 2, "target": 5, "relation": "includes"},
    {"source": 2, "target": 6, "relation": "includes"},
    {"source": 2, "target": 7, "relation": "includes"},
    
    {"source": 3, "target": 8, "relation": "includes"},
    {"source": 3, "target": 9, "relation": "includes"},
    {"source": 3, "target": 10, "relation": "includes"},
    
    {"source": 4, "target": 11, "relation": "includes"},
    {"source": 4, "target": 12, "relation": "includes"},
]

    # 运行函数
    export_knowledge_graph_to_excel(nodes, edges, output_path="f1.xlsx")
